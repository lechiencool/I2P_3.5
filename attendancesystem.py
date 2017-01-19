from attendancesysclassdata import *
import datetime
from openpyxl import Workbook

present = []
tardy = []
absent = authorized[:]

def reset(present, tardy):

    global absent
    todaydate = '{:%b %d %Y}'.format(datetime.datetime.now())

    # Clearing present/tardy list from day before & resetting absent list
    if len(present) == 0:
        if len(tardy) == 0:
            process(todaydate)
        else:
            if len(present) > 0:
                if todaydate != present[0][1]:
                    present[:] = []
                    tardy[:] = []
                    absent = authorized[:]
                    process(todaydate)
                else:
                    process(todaydate)
            elif len(tardy) > 0:
                if todaydate != tardy[0][1]:
                    present[:] = []
                    tardy[:] = []
                    absent = authorized[:]
                    process(todaydate)
                else:
                    process(todaydate)
    else:
        if len(present) > 0:
            if todaydate != present[0][1]:
                present[:] = []
                tardy[:] = []
                absent = authorized[:]
                process(todaydate)
            else:
                process(todaydate)
        elif len(tardy) > 0:
            if todaydate != tardy[0][1]:
                present[:] = []
                tardy[:] = []
                absent = authorized[:]
                process(todaydate)
            else:
                process(todaydate)

def process(todaydate):
    # Start of verification process
    scan = input("Please scan your card! ")

    # Checking if dates match
    datecheck = '{:%b %d %Y}'.format(datetime.datetime.now())
    if todaydate != datecheck:
        print("\nSorry, there was an error. Please try again!\n")
        reset(present, tardy)
    else:
        # Checking if RFID & individual is registered
        if rfidnamematcher.count(scan) == 0:
            print("\nNot authorized. Sorry!\n")
            reset(present, tardy)
        else:
            findauthorized = rfidnamematcher[rfidnamematcher.index(scan) + 1]
            authorizationcheck(findauthorized, authorized, todaydate)

def authorizationcheck(findauthorized, authorized, todaydate):
    # Checking if individual is on "authorized" list
    if authorized.count(findauthorized) == 0:
        print("\nNot authorized. Sorry!\n")
        reset(present, tardy)
    else:
        doublesignincheck(present, tardy, findauthorized, todaydate)

def doublesignincheck(present, tardy, findauthorized, todaydate):
    if present.count([findauthorized, todaydate]) == 0:
        if tardy.count([findauthorized, todaydate]) == 0:
            # Note this is a placeholder, want to create a second authentication step before moving onto tardycheck()
            tardycheck(findauthorized, timeofclass, present, tardy, todaydate)
        else:
            print("\nYou have already signed in!\n")
            reset(present, tardy)
    else:
        print("\nYou have already signed in!\n")
        reset(present, tardy)

def tardycheck(findauthorized, timeofclass, present, tardy, todaydate):
    timenow = '{:%H:%M:%S}'.format(datetime.datetime.now())

    # Checking if individual is tardy or on time
    if timenow > timeofclass:
        print("\nTime now:", timenow)
        print("Class starts:", timeofclass)
        print("Welcome to class,", findauthorized,"! You are late and have been marked as tardy. \n")
        tardy.append([findauthorized,todaydate])
        absent.remove(findauthorized)
        record(present, tardy, todaydate)
    else:
        print("\nTime now:", timenow)
        print("Class starts:", timeofclass)
        print("Welcome to class,", findauthorized,"! You have been marked as present. \n")
        present.append([findauthorized, todaydate])
        absent.remove(findauthorized)
        record(present, tardy, todaydate)

def record(present, tardy, todaydate):
    global absent

    # Creating record list from list of lists
    rectardy = [item[0] for item in tardy]
    recpresent = [item[0] for item in present]

    # Defining workbook and worksheet name
    wb = Workbook()
    dest_filename = todaydate + ".xlsx"
    ws1 = wb.active
    ws1.title = "Attendance for " + todaydate

    # Writing to the workbook
    ws1['A1'] = "Present"
    ws1['B1'] = "Tardy"
    ws1['C1'] = "Absent"

    setboldp = ws1['A1']
    setboldp.font = setboldp.font.copy(bold=True)
    setboldt = ws1['B1']
    setboldt.font = setboldt.font.copy(bold=True)
    setbolda = ws1['C1']
    setbolda.font = setbolda.font.copy(bold=True)

    rp = 2
    rt = 2
    ra = 2
    for p in recpresent:
        ws1.cell(row=rp, column=1).value = p
        rp += 1

    for t in rectardy:
        ws1.cell(row=rt, column=2).value = t
        rt += 1

    for a in absent:
        ws1.cell(row=ra, column=3).value = a
        ra += 1

    # Setting column widths
    ws1.column_dimensions["A"].width = 30.0
    ws1.column_dimensions["B"].width = 30.0
    ws1.column_dimensions["C"].width = 30.0

    wb.save(filename = dest_filename)

    reset(present, tardy)

reset(present, tardy)
